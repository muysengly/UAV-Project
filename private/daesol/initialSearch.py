import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import matplotlib.patches as mpatches
import openpyxl as xl
import pandas as pd
from scipy.spatial import distance_matrix

def makeUAV(xpos,ypos,maxbeam,uavno):
    uav_x_pos = xpos  # x position of uav 0,1,...,GRID_SIZE-1
    uav_y_pos = ypos  # y position of uav 0,1,...,GRID_SIZE-1
    uav_number = "UAV"+str(uavno)
    uav_x = GRID[1, uav_x_pos, uav_y_pos]
    uav_y = GRID[0, uav_x_pos, uav_y_pos]
    uav_z2 = UAV_ALTITUDE  # uav's altitude [meter]

    plt.scatter(x=uav_x, y=uav_y, c="green")
    plt.text(x=uav_x - 3.5, y=uav_y - 4, s=uav_number)

def makeBeamCircle(xpos,ypos,maxbeam,color):
    uav_x_pos = xpos  # x position of uav 0,1,...,GRID_SIZE-1
    uav_y_pos = ypos  # y position of uav 0,1,...,GRID_SIZE-1
    uav_x = GRID[1, uav_x_pos, uav_y_pos]
    uav_y = GRID[0, uav_x_pos, uav_y_pos]
    uav_z2 = UAV_ALTITUDE  # uav's altitude [meter]

    beam_circle = Ellipse(
        xy=(uav_x, uav_y),
        width=maxbeam,
        height=maxbeam,
        angle=0,
        edgecolor="none",
        facecolor=color,
        alpha=0.2,
    )
    uav_beam = ax.add_patch(beam_circle)

def searchGu(x,y):
    uavx=(x+0.5)*10
    uavy=(y+0.5)*10
    tmp_count=0
    count=0
    for k in range(10):
        if uavx-17<=gu_x[k]<=uavx+17 and uavy-17<=gu_y[k]<=uavy+17:
            if gu_memory[0][k]!=gu_x[k] and gu_memory[1][k]!=gu_y[k]:
                print("gu #: ",end='')
                counter1[0]+=1
                print(k)
                guno[0]=k
                gu_memory[0][k]=gu_x[k]
                gu_memory[1][k]=gu_y[k]
            else:
                continue
        
#intial parameter
NUM_GU = 10  # number of ground users

X_MIN = 0  # minimum x-axis [meter]
X_MAX = 100  # maximum x-axis [meter]
Y_MIN = 0  # minimum y-axis [meter]
Y_MAX = 100  # maximum y-axis [mseter]

UAV_ALTITUDE = 10  # altitude of uav [meter]
MAX_BEAM_ANGLE = 60  # maximum beamforming angle [degree]

# maximum beamforming diameter [meter]
MAX_BEAM_DIAMETER = 2*UAV_ALTITUDE*np.tan(MAX_BEAM_ANGLE*np.pi/180)


X_GRID = 10  # number of x grid
Y_GRID = 10  # number of y grid


UAV_TX_POWER = 30  # uav's transmit power in [dBm]

#initial variables
t = 0  # time [seconds]

gu_bat = np.zeros((NUM_GU,)) # battery of ground user [mWh]

#generate variables
# generate ground user location randomly x,y,z [meters]
gu_x = np.random.uniform(low=X_MIN, high=X_MAX, size=(NUM_GU,))
gu_y = np.random.uniform(low=Y_MIN, high=Y_MAX, size=(NUM_GU,))
gu_z = np.zeros((NUM_GU,))

# print
gu_x, gu_y, gu_z

# generate meshgrid
tmp_x = np.linspace(X_MIN + X_GRID/2,X_MAX-X_GRID/2,X_GRID)
tmp_y = np.linspace(Y_MIN + Y_GRID/2,Y_MAX-Y_GRID/2,Y_GRID)
GRID = np.array(np.meshgrid(tmp_x, tmp_y))

GRID

# generate uav location

uav_x_pos = 1  # x position of uav 0,1,...,GRID_SIZE-1
uav_y_pos = 0  # y position of uav 0,1,...,GRID_SIZE-1

uav_x = GRID[1, uav_x_pos, uav_y_pos]
uav_y = GRID[0, uav_x_pos, uav_y_pos]
uav_z = UAV_ALTITUDE  # uav's altitude [meter]

# print
uav_x, uav_y, uav_z
fig, ax = plt.subplots(figsize=(6, 6))

for i in range(NUM_GU):
    plt.scatter(x=gu_x[i], y=gu_y[i], c="blue")
    plt.text(x=gu_x[i] - 3.5, y=gu_y[i] - 4, s=f"GU-{i}")
    plt.text(x=gu_x[i] - 6, y=gu_y[i] - 7, s=f"{gu_bat[i]}mWh")

print(MAX_BEAM_DIAMETER)
print(gu_x)
print(gu_y)
print(gu_z)
gu_memory=np.ones((2,10)) + 10
tmp_state=np.ones((10,2)) 

countUAV=1
counter1=np.zeros(1)
guno=np.zeros(1)
for i in range(1,11,2):
    if i==1 or i==5 or i==9:
        for k in range(0,10,2):
            if counter1[0]!=10:
                searchGu(i,k)
                makeUAV(i,k,MAX_BEAM_DIAMETER,countUAV)
                if countUAV%2 == 0:
                    makeBeamCircle(i,k,MAX_BEAM_DIAMETER,"yellow")
                else:
                    makeBeamCircle(i,k,MAX_BEAM_DIAMETER,"orange")
                if countUAV%5 !=0 and counter1[0]!=10:
                    plt.plot([(i+0.5)*10,(i+0.5)*10],[(k+0.5)*10,(k+2.5)*10],color="green")
                elif countUAV==25:
                    searchGu(i,k)
                  
                else:
                    if counter1[0]!=10:
                        plt.plot([(i+0.5)*10,(i+2.5)*10],[(k+0.5)*10,(k+0.5)*10],color="green")
                
                print(countUAV)
                countUAV+=1
            else:
                break
          
            
    else:
        for k in range(8,-2,-2):
            if counter1[0]!=10:
                searchGu(i,k)
                makeUAV(i,k,MAX_BEAM_DIAMETER,countUAV)
                if countUAV%2 == 0:
                    makeBeamCircle(i,k,MAX_BEAM_DIAMETER,"yellow")
                else:
                    makeBeamCircle(i,k,MAX_BEAM_DIAMETER,"orange")
                if countUAV%5 ==0 and counter1[0]!=10:    
                    plt.plot([(i+0.5)*10,(i+2.5)*10],[(k+0.5)*10,(k+0.5)*10],color="green")
                else:
                    if counter1[0]!=10:
                        plt.plot([(i+0.5)*10,(i+0.5)*10],[(k+0.5)*10,(k-1.5)*10],color="green")
                print(countUAV)
                countUAV+=1
            else:
                break
      
            
print(gu_memory)
print(counter1)
print(guno[0])

wb=xl.Workbook()
sheet1= wb.active
sheet1.title='location'
sheet1.cell(row=1,column=1,value='GU location')
sheet1.cell(row=2,column=1,value='GU x-loc')
sheet1.cell(row=2,column=2,value='GU y-loc')
sheet1.cell(row=2,column=3,value=guno[0])
for x in range(2):
    for y in range(10):
        sheet1.cell(row=3+y, column=1+x,value=gu_memory[x][y])
wb.save(filename='locationInformation.xlsx')

#pd.set_option('display.max_columns',None)
df=pd.read_excel('locationInformation.xlsx')
print(df)

plt.xlabel("x-axis [m]")
plt.ylabel("y-axis [m]")
plt.title("Simulation Environment")
plt.xticks(np.arange(X_MIN, X_MAX + 1, X_GRID))
plt.yticks(np.arange(Y_MIN, Y_MAX + 1, Y_GRID))
plt.xlim(X_MIN, X_MAX)
plt.ylim(Y_MIN, Y_MAX)
plt.grid()
plt.show()
