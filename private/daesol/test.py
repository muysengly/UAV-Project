import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import matplotlib.patches as mpatches
import openpyxl as xl
import pandas as pd
from scipy.spatial import distance_matrix
import math

def makeUAV(xpos,ypos,maxbeam,uavno):
    uav_x_pos = xpos  # x position of uav 0,1,...,GRID_SIZE-1
    uav_y_pos = ypos  # y position of uav 0,1,...,GRID_SIZE-1
    uav_number = "UAV"+str(uavno)
    uav_x = GRID[1, uav_x_pos, uav_y_pos]
    uav_y = GRID[0, uav_x_pos, uav_y_pos]
    uav_z2 = UAV_ALTITUDE  # uav's altitude [meter]

    plt.scatter(x=uav_x, y=uav_y, c="green")
    plt.text(x=uav_x - 3.5, y=uav_y - 4, s=uav_number)

def makeBeamCircle(xpos,ypos,maxbeam,color):
    beam_circle = Ellipse(
        xy=(xpos, ypos),
        width=maxbeam,
        height=maxbeam,
        angle=0,
        edgecolor="none",
        facecolor=color,
        alpha=0.2,
    )
    uav_beam = ax.add_patch(beam_circle)

#intial parameter
NUM_GU = 10  # number of ground users

X_MIN = 0  # minimum x-axis [meter]
X_MAX = 100  # maximum x-axis [meter]
Y_MIN = 0  # minimum y-axis [meter]
Y_MAX = 100  # maximum y-axis [mseter]

UAV_ALTITUDE = 10  # altitude of uav [meter]
MAX_BEAM_ANGLE = 60  # maximum beamforming angle [degree]

# maximum beamforming diameter [meter]
MAX_BEAM_DIAMETER = 2*UAV_ALTITUDE*np.tan(MAX_BEAM_ANGLE*np.pi/180)


X_GRID = 10  # number of x grid
Y_GRID = 10  # number of y grid


UAV_TX_POWER = 30  # uav's transmit power in [dBm]

#initial variables
t = 0  # time [seconds]

gu_bat = np.zeros((NUM_GU,)) # battery of ground user [mWh]

#generate variables
# generate ground user location randomly x,y,z [meters]
gu_x = np.random.uniform(low=X_MIN, high=X_MAX, size=(NUM_GU,))
gu_y = np.random.uniform(low=Y_MIN, high=Y_MAX, size=(NUM_GU,))
gu_z = np.zeros((NUM_GU,))

# print
gu_x, gu_y, gu_z

# generate meshgrid
tmp_x = np.linspace(X_MIN + X_GRID/2,X_MAX-X_GRID/2,X_GRID)
tmp_y = np.linspace(Y_MIN + Y_GRID/2,Y_MAX-Y_GRID/2,Y_GRID)
GRID = np.array(np.meshgrid(tmp_x, tmp_y))

GRID

# generate uav location

uav_x_pos = 1  # x position of uav 0,1,...,GRID_SIZE-1
uav_y_pos = 0  # y position of uav 0,1,...,GRID_SIZE-1

uav_x = GRID[1, uav_x_pos, uav_y_pos]
uav_y = GRID[0, uav_x_pos, uav_y_pos]
uav_z = UAV_ALTITUDE  # uav's altitude [meter]

# print
uav_x, uav_y, uav_z
fig, ax = plt.subplots(figsize=(6, 6))

for i in range(NUM_GU):
    plt.scatter(x=gu_x[i], y=gu_y[i], c="blue")
    plt.text(x=gu_x[i] - 3.5, y=gu_y[i] - 4, s=f"GU-{i}")
    plt.text(x=gu_x[i] - 6, y=gu_y[i] - 7, s=f"{gu_bat[i]}mWh")
makeBeamCircle(10,20,MAX_BEAM_DIAMETER/2,'yellow')

"""
wb=xl.Workbook()
sheet1= wb.active
sheet1.title='location'
sheet1.cell(row=1,column=1,value='GU location')
sheet1.cell(row=2,column=1,value='GU x-loc')
sheet1.cell(row=2,column=2,value='GU y-loc')
sheet1.cell(row=2,column=3,value=guno[0])
for x in range(2):
    for y in range(10):
        sheet1.cell(row=3+y, column=1+x,value=gu_memory[x][y])
wb.save(filename='locationInformation.xlsx')

#pd.set_option('display.max_columns',None)
df=pd.read_excel('locationInformation.xlsx')
print(df)
"""
plt.xlabel("x-axis [m]")
plt.ylabel("y-axis [m]")
plt.title("Simulation Environment")
plt.xticks(np.arange(X_MIN, X_MAX + 1, X_GRID))
plt.yticks(np.arange(Y_MIN, Y_MAX + 1, Y_GRID))
plt.xlim(X_MIN, X_MAX)
plt.ylim(Y_MIN, Y_MAX)
plt.grid()
plt.show()
