import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import matplotlib.patches as mpatches
import openpyxl as xl


def searchGu(x,y):

    tmp_count=0
    count=0
    for k in range(10):
        if x-17<=gu_x[k]<=x+17 and y-17<=gu_y[k]<=y+17:
            if gu_memory[0][k]!=gu_x[k] and gu_memory[1][k]!=gu_y[k]:
                counter1[0]+=1
                print(k)
                guno[0]=k
                gu_memory[0][k]=gu_x[k]
                gu_memory[1][k]=gu_y[k]
            else:
                continue
def makeBeamCircle(xpos,ypos,maxbeam,color):
    beam_circle = Ellipse(
        xy=(xpos, ypos),
        width=maxbeam,
        height=maxbeam,
        angle=0,
        edgecolor="none",
        facecolor=color,
        alpha=0.2,
    )
    uav_beam = ax.add_patch(beam_circle)
NUM_GU = 10  # number of ground users

X_MIN = 0  # minimum x-axis [meter]
X_MAX = 100  # maximum x-axis [meter]
Y_MIN = 0  # minimum y-axis [meter]
Y_MAX = 100  # maximum y-axis [mseter]

UAV_ALTITUDE = 10  # altitude of uav [meter]
MAX_BEAM_ANGLE = 60  # maximum beamforming angle [degree]

# maximum beamforming diameter [meter]
MAX_BEAM_DIAMETER = 2*UAV_ALTITUDE*np.tan(MAX_BEAM_ANGLE*np.pi/180)


X_GRID = 10  # number of x grid
Y_GRID = 10  # number of y grid


UAV_TX_POWER = 30  # uav's transmit power in [dBm]

#initial variables
t = 0  # time [seconds]

gu_bat = np.zeros((NUM_GU,)) # battery of ground user [mWh]

#generate variables
# generate ground user location randomly x,y,z [meters]
gu_x = np.random.uniform(low=X_MIN, high=X_MAX, size=(NUM_GU,))
gu_y = np.random.uniform(low=Y_MIN, high=Y_MAX, size=(NUM_GU,))
gu_z = np.zeros((NUM_GU,))

# print
gu_x, gu_y, gu_z

# generate meshgrid
tmp_x = np.linspace(X_MIN + X_GRID/2,X_MAX-X_GRID/2,X_GRID)
tmp_y = np.linspace(Y_MIN + Y_GRID/2,Y_MAX-Y_GRID/2,Y_GRID)
GRID = np.array(np.meshgrid(tmp_x, tmp_y))

GRID

# generate uav location

uav_x_pos = 1  # x position of uav 0,1,...,GRID_SIZE-1
uav_y_pos = 0  # y position of uav 0,1,...,GRID_SIZE-1

uav_x = GRID[1, uav_x_pos, uav_y_pos]
uav_y = GRID[0, uav_x_pos, uav_y_pos]
uav_z = UAV_ALTITUDE  # uav's altitude [meter]
uav_x, uav_y, uav_z
fig, ax = plt.subplots(figsize=(6, 6))
gu_memory=np.ones((2,10)) + 10
guno=np.zeros(1)

for i in range(NUM_GU):
    plt.scatter(x=gu_x[i], y=gu_y[i], c="blue")
    plt.text(x=gu_x[i] - 3.5, y=gu_y[i] - 4, s=f"GU-{i}")
    plt.text(x=gu_x[i] - 6, y=gu_y[i] - 7, s=f"{gu_bat[i]}mWh")

counter1=np.zeros(1)
points_opt=[[32.1181633, 88.8492449],
 [65.3537403, 88.8089527],
 [54.9709193, 34.4695308],
 [82.1278111, 60.1633634],
 [17.9589312, 34.9669854],
 [8.68285224, 8.00136375],
 [81.4411862, 7.45478934],
 [98.8207531, 89.2055321],
 [44.9338381, 8.96557517],
 [1.45349405, 95.4804783],
 [91.3545575, 33.1080821],
 [46.0975763, 59.9639461],
 [9.93585808, 61.3039564]]
route1=[9,0,1,7,3,10,6,8,5,4,12,11,2]
uavnum=1
count2=1
for i in range(13):
    plt.scatter(x=points_opt[route1[i]][0], y=points_opt[route1[i]][1], c="red")
    ax.text(x=points_opt[route1[i]][0] - 3.5, y=points_opt[route1[i]][1] - 4, s=f"UAV-{i}")
    if count2%2==0:
        makeBeamCircle(points_opt[route1[i]][0], points_opt[route1[i]][1],40,'yellow')
    else:
        makeBeamCircle(points_opt[route1[i]][0], points_opt[route1[i]][1],40,'blue')
    searchGu(points_opt[route1[i]][0],points_opt[route1[i]][1])
    if counter1[0]==10:
        break
    count2+=1
plt.xlabel("x-axis [m]")
plt.ylabel("y-axis [m]")
plt.title("Simulation Environment")
plt.xticks(np.arange(X_MIN, X_MAX + 1, X_GRID))
plt.yticks(np.arange(Y_MIN, Y_MAX + 1, Y_GRID))
plt.xlim(X_MIN, X_MAX)
plt.ylim(Y_MIN, Y_MAX)
plt.grid()
plt.show()
wb=xl.Workbook()
sheet1= wb.active
sheet1.title='location'
sheet1.cell(row=1,column=1,value='GU location')
sheet1.cell(row=2,column=1,value='GU x-loc')
sheet1.cell(row=2,column=2,value='GU y-loc')
sheet1.cell(row=2,column=3,value=guno[0])
for x in range(2):
    for y in range(10):
        sheet1.cell(row=3+y, column=1+x,value=gu_memory[x][y])
wb.save(filename='locInfo.xlsx')



